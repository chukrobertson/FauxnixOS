from __future__ import annotations

from pathlib import Path
import csv

from fauxnix_tools.config import config
from fauxnix_tools.utils import guess_mime

pytesseract = None
fitz = None
Document = None
load_workbook = None
Image = None


def _lazy_tesseract():
    global pytesseract
    if pytesseract is None:
        import pytesseract as pt
        pt.pytesseract.tesseract_cmd = config.tesseract_cmd
        pytesseract = pt
    return pytesseract


def _lazy_pil():
    global Image
    if Image is None:
        from PIL import Image as PILImage
        Image = PILImage
    return Image


def extract_pdf_text(path: Path) -> str:
    global fitz
    try:
        if fitz is None:
            import fitz as f
            fitz = f
        doc = fitz.open(path)
        return "\n".join(page.get_text("text") for page in doc).strip()
    except Exception as e:
        return f"[PDF extraction error] {e}"


def extract_docx_text(path: Path) -> str:
    global Document
    try:
        if Document is None:
            from docx import Document as Doc
            Document = Doc
        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text).strip()
    except Exception as e:
        return f"[DOCX extraction error] {e}"


def extract_xlsx_text(path: Path, max_cells: int = 4000) -> str:
    global load_workbook
    try:
        if load_workbook is None:
            from openpyxl import load_workbook as lw
            load_workbook = lw
        wb = load_workbook(path, read_only=True, data_only=True)
        lines = []
        count = 0
        for ws in wb.worksheets:
            lines.append(f"## Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                vals = [str(v) for v in row if v is not None and str(v).strip()]
                if vals:
                    lines.append(" | ".join(vals))
                    count += len(vals)
                if count >= max_cells:
                    lines.append("[TRUNCATED]")
                    return "\n".join(lines)
        return "\n".join(lines).strip()
    except Exception as e:
        return f"[XLSX extraction error] {e}"


def extract_csv_text(path: Path, max_lines: int = 1000) -> str:
    try:
        lines = []
        with open(path, "r", encoding="utf-8", errors="ignore", newline="") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= max_lines:
                    lines.append("[TRUNCATED]")
                    break
                lines.append(" | ".join(row))
        return "\n".join(lines).strip()
    except Exception as e:
        return f"[CSV extraction error] {e}"


def extract_txt_text(path: Path, max_chars: int = 200000) -> str:
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read(max_chars).strip()
    except Exception as e:
        return f"[TXT extraction error] {e}"


def extract_image_ocr(path: Path) -> str:
    try:
        pt = _lazy_tesseract()
        pil = _lazy_pil()
        return pt.image_to_string(pil.open(path)).strip()
    except Exception as e:
        return f"[OCR extraction error] {e}"


EXTRACTOR_MAP = {
    ".pdf": extract_pdf_text,
    ".docx": extract_docx_text,
    ".xlsx": extract_xlsx_text,
    ".csv": extract_csv_text,
    ".txt": extract_txt_text,
    ".md": extract_txt_text,
    ".rtf": extract_txt_text,
    ".json": extract_txt_text,
    ".xml": extract_txt_text,
    ".log": extract_txt_text,
    ".yaml": extract_txt_text,
    ".yml": extract_txt_text,
    ".toml": extract_txt_text,
    ".ini": extract_txt_text,
    ".cfg": extract_txt_text,
    ".conf": extract_txt_text,
}


def extract_any(path: Path) -> str:
    ext = path.suffix.lower()
    mime = guess_mime(path)
    extractor = EXTRACTOR_MAP.get(ext)
    if extractor:
        return extractor(path)
    if mime.startswith("image/"):
        return extract_image_ocr(path)
    return ""


def extract_text(path: Path | str, format: str = "auto", max_chars: int = 10000) -> str:
    p = Path(path)
    fmt = format.lower()

    if fmt == "auto" or fmt == "all":
        return extract_any(p)

    dispatcher = {
        "pdf": extract_pdf_text,
        "docx": extract_docx_text,
        "xlsx": lambda p: extract_xlsx_text(p, max_cells=max_chars),
        "csv": lambda p: extract_csv_text(p, max_lines=max(max_chars // 100, 100)),
        "txt": lambda p: extract_txt_text(p, max_chars=max_chars),
        "ocr": extract_image_ocr,
    }

    fn = dispatcher.get(fmt, extract_any)
    return fn(p)
