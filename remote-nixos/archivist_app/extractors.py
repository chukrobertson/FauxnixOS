from pathlib import Path
import csv
import fitz
import pytesseract
from PIL import Image
from docx import Document
from openpyxl import load_workbook

from app.config import TESSERACT_CMD
from app.utils import guess_mime

pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

PLAIN_TEXT_EXTS: set[str] = {
    ".txt", ".md", ".rst", ".rtf", ".json", ".xml", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".conf",
    ".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".htm", ".css", ".scss", ".less",
    ".c", ".h", ".cpp", ".hpp", ".cc", ".cxx", ".cs", ".java", ".kt", ".swift",
    ".go", ".rs", ".rb", ".php", ".pl", ".pm", ".sh", ".bash", ".zsh", ".fish",
    ".ps1", ".bat", ".cmd", ".vbs", ".sql", ".ipynb", ".proto", ".f", ".f90",
    ".R", ".r", ".scala", ".clj", ".elm", ".ex", ".exs", ".erl", ".hrl",
    ".lua", ".nim", ".pas", ".dart", ".gradle", ".sbt", ".tf", ".hcl",
    ".dockerfile", ".makefile", ".cmake", ".m", ".mm",
    ".tex", ".bib", ".cls", ".sty",
    ".svelte", ".vue", ".astro",
}

PLAIN_TEXT_NAMES: set[str] = {
    "makefile", "dockerfile", "docker-compose.yml", "docker-compose.yaml",
    "gemfile", "rakefile", "procfile", "requirements.txt",
    ".gitignore", ".gitattributes", ".editorconfig", ".env",
    "cargo.toml", "package.json", "tsconfig.json", "webpack.config.js",
}


def supports_text_extraction(path: Path) -> bool:
    return path.suffix.lower() in PLAIN_TEXT_EXTS or path.name.lower() in PLAIN_TEXT_NAMES


def extract_pdf_text(path: Path) -> str:
    try:
        doc = fitz.open(path)
        return "\n".join(page.get_text("text") for page in doc).strip()
    except Exception as e:
        return f"[PDF extraction error] {e}"


def extract_docx_text(path: Path) -> str:
    try:
        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text).strip()
    except Exception as e:
        return f"[DOCX extraction error] {e}"


def extract_xlsx_text(path: Path, max_cells: int = 4000) -> str:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        lines = []
        count = 0
        for ws in wb.worksheets:
            lines.append(f"## Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                vals = [str(v) for v in row if v is not None and str(v).strip()]
                if vals:
                    lines.append(" | ".join(vals))
                    count += len(vals)
                if count >= max_cells:
                    lines.append("[TRUNCATED]")
                    return "\n".join(lines)
        return "\n".join(lines).strip()
    except Exception as e:
        return f"[XLSX extraction error] {e}"


def extract_csv_text(path: Path, max_lines: int = 1000) -> str:
    try:
        lines = []
        with open(path, "r", encoding="utf-8", errors="ignore", newline="") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= max_lines:
                    lines.append("[TRUNCATED]")
                    break
                lines.append(" | ".join(row))
        return "\n".join(lines).strip()
    except Exception as e:
        return f"[CSV extraction error] {e}"


def extract_txt_text(path: Path, max_chars: int = 200000) -> str:
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read(max_chars).strip()
    except Exception as e:
        return f"[TXT extraction error] {e}"


def extract_image_ocr(path: Path) -> str:
    try:
        return pytesseract.image_to_string(Image.open(path)).strip()
    except Exception as e:
        return f"[OCR extraction error] {e}"


def extract_any(path: Path) -> str:
    ext = path.suffix.lower()
    mime = guess_mime(path)
    if ext == ".pdf":
        return extract_pdf_text(path)
    if ext == ".docx":
        return extract_docx_text(path)
    if ext == ".xlsx":
        return extract_xlsx_text(path)
    if ext == ".csv":
        return extract_csv_text(path)
    if ext in [".txt", ".md", ".rtf", ".json", ".xml", ".log"]:
        return extract_txt_text(path)
    if mime.startswith("image/"):
        return extract_image_ocr(path)
    return ""
